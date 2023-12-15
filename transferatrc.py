from openpyxl import load_workbook

# Load the source Excel file
source_file_path = r"C:\Users\divya\Documents\ETL extracted\ATRC extracted.xlsx"
source_workbook = load_workbook(source_file_path)
source_sheet = source_workbook['Sheet1']  # Change 'Sheet1' to your source sheet name

# Create a new Excel file or load an existing one
output_file_path = r"C:\Users\divya\Documents\ETL extracted\ATRC transfer.xlsx"
output_workbook = load_workbook(output_file_path)

# Create a new sheet in the output file
output_sheet = output_workbook.create_sheet(title='Sheet2')  # Change 'Sheet2' to your desired destination sheet name

# Copy data from source sheet to destination sheet
for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row, values_only=True):
    output_sheet.append(row)

# Save the changes to the output file
output_workbook.save(output_file_path)

print("Data transfer complete.")
