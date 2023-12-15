import openpyxl

def extract_and_save_data(source_path, destination_path):
    try:
        # Load the source workbook
        print(f"Loading the source workbook: {source_path}")
        source_workbook = openpyxl.load_workbook(source_path)
        print("Source workbook loaded successfully.")

        # Create a new workbook for the extracted data
        destination_workbook = openpyxl.Workbook()
        destination_sheet = destination_workbook.active

        # Extract data from the source sheet and append it to the destination sheet
        for row_num, row in enumerate(source_workbook.active.iter_rows(values_only=True), start=1):
            print(f"Extracting data from row {row_num}: {row}")
            destination_sheet.append(row)

        # Save the destination workbook to the specified path
        print(f"Saving the destination workbook: {destination_path}")
        destination_workbook.save(destination_path)
        print("Destination workbook saved successfully.")

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Close both workbooks
        if 'source_workbook' in locals():
            source_workbook.close()

        if 'destination_workbook' in locals():
            destination_workbook.close()

# Specify the paths for the source and destination Excel files
source_excel_file = r"C:\Users\divya\Documents\ETL extracted\Copy of Copy_of_ATRC_-_EndNov2018_Send(2)(1).xlsx"
output_excel_file = r"C:\Users\divya\Documents\ETL extracted\ATRC extracted.xlsx"

# Call the function to extract and save data
extract_and_save_data(source_excel_file, output_excel_file)


